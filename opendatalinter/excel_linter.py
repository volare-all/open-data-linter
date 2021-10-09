import csv
import datetime
import io

import openpyxl
from openpyxl.cell import Cell

from .csv_linter import CSVLinter
from .funcs import before_check_1_1
from .vo import LintResult


def ws2csv(ws) -> str:
    with io.StringIO() as s:
        writer = csv.writer(s)
        for row in ws.rows:
            writer.writerow([__to_value(cell) for cell in row])
        return s.getvalue()


def __to_value(cell: Cell):
    value = cell.value
    # 日付データを数値データに変換し、無用なエラーを抑制する
    if isinstance(value, datetime.datetime):
        return value.timestamp()
    elif isinstance(value, datetime.date):
        datetime.datetime.combine(value, datetime.time()).timestamp()
    elif isinstance(value, datetime.time):
        return (value.hour * 60 + value.minute) * 60 + value.second
    else:
        return cell.value


class ExcelLinter:
    def __getattr__(self, name):
        return getattr(self.csv_linter, name)

    def __init__(self,
                 data: bytes,
                 filename: str,
                 title_line_num=None,
                 header_line_num=None):
        with io.BytesIO(data) as f:
            wb = openpyxl.load_workbook(f)
            # df = pd.read_excel(f, header=None)

        # ToDo: 複数のシートに対応できるように
        for sheetname in wb.sheetnames:
            self.ws = wb[sheetname]
            self.text = ws2csv(self.ws)
            break

        self.wb = wb
        self.csv_linter = CSVLinter(self.text.encode(),
                                    "from_excel.csv",
                                    title_line_num=title_line_num,
                                    header_line_num=header_line_num)

    @before_check_1_1
    def check_1_4(self):
        """チェック項目1-4に沿って、セルの結合をしていないか確認する。
        """
        invalid_cells = []
        for merged_cell in self.ws.merged_cells:
            b = merged_cell.bounds
            invalid_cells.append((b[1] - 1, b[0] - 1))  # 0-base-index
        return LintResult.gen_single_error_message_result(
            "結合されたセルが存在します", invalid_cells)

    @before_check_1_1
    def check_1_7(self):
        """チェック項目1-7に沿って、数式を使⽤している場合は数値データに修正しているか確認する。

        Note:
            '='から始まるセルを invalid とみなす。
        """
        invalid_cells = []
        for r in range(0, self.ws.max_row):
            for c in range(0, self.ws.max_column):
                if str(self.ws.cell(r + 1, c + 1).value).startswith("="):
                    invalid_cells.append((r, c))
        return LintResult.gen_single_error_message_result(
            "数式が含まれています", invalid_cells)
